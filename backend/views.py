from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, authenticate
from .models import Contract, Client, Invoice, Expense, Party
from .forms import ContractForm, ClientForm, InvoiceForm, ExpenseForm, PartyForm
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import logging

logger = logging.getLogger(__name__)



def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Log the user in after signing up
            return redirect('dashboard')  # Redirect to the dashboard after signup
    else:
        form = UserCreationForm()
    return render(request, 'backend/signup.html', {'form': form})

def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')  # Redirect to the dashboard after login
    else:
        form = AuthenticationForm()
    return render(request, 'backend/login.html', {'form': form})

def logout_view(request):
    request.user.auth_backend.logout(request)
    return redirect('home')

@login_required
def dashboard(request):
    total_contracts = Contract.objects.count()
    pending_invoices = Contract.objects.filter(invoice__isnull=True).count()
    total_clients = Client.objects.count()
    
    # From contract Model 
    contract_net_invoice = Contract.objects.aggregate(Sum('net_invoice'))['net_invoice__sum']

    context = {
        'total_contracts': total_contracts,
        'total_clients': total_clients,
        'pending_invoices': pending_invoices,
        'contract_net_invoice': contract_net_invoice
    }
    
    return render(request, 'backend/dashboard.html', context=context)

@login_required
@login_required
def contract_list(request):
    contracts = Contract.objects.all()

    if request.method == 'POST' and 'export' in request.POST:
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define the column headers
        column_headers = ['Contractor Name', 'Provider Name', 'Start Date', 'End Date', 'Net Invoice']

        # Write the column headers to the worksheet
        for i, header in enumerate(column_headers, start=1):
            ws[f'A{i}'] = header

        # Write the data to the worksheet
        for i, contract in enumerate(contracts, start=2):
            ws[f'A{i}'] = contract.contractor
            ws[f'B{i}'] = contract.provider
            ws[f'C{i}'] = contract.practice_name.start_date
            ws[f'D{i}'] = contract.practice_name.end_date
            ws[f'E{i}'] = contract.net_invoice

        # Create an HTTP response with the workbook as the content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=contracts.xlsx'
        wb.save(response)
        return response

    return render(request, 'backend/contract_list.html', {'contracts': contracts})
@login_required
@login_required
def contract_create(request):
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            contract = form.save()
            return redirect('contract_list')
        else:
            logger.error(f"Form errors: {form.errors}")
    else:
        form = ContractForm()

    return render(request, 'backend/contract_form.html', {
        'form': form,
        'contract': None,  # Add this line
        'calculated_percentage': 0,
        'total_work_amount': 0,
        'total_transactions_amount': 0
    })
@login_required
def contract_update(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            form.save()
            return redirect('contract_list')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'backend/contract_form.html', {'form': form, 'contract': contract})

@login_required
def contract_delete(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('contract_list')
    return render(request, 'backend/contract_confirm_delete.html', {'contract': contract})

@login_required
def client_list(request):
    clients = Client.objects.all()

    if request.method == 'POST' and 'export' in request.POST:
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define the column headers
        column_headers = ['Client Name', 'Address', 'Phone Number', 'Email']

        # Write the column headers to the worksheet
        for i, header in enumerate(column_headers, start=1):
            ws[f'A{i}'] = header

        # Write the data to the worksheet
        for i, client in enumerate(clients, start=2):
            ws[f'A{i}'] = client.practice_name
            ws[f'B{i}'] = client.address
            ws[f'C{i}'] = client.phone_number
            ws[f'D{i}'] = client.email

        # Create an HTTP response with the workbook as the content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=clients.xlsx'
        wb.save(response)
        return response

    return render(request, 'backend/client_list.html', {'clients': clients})

@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientForm()
    return render(request, 'backend/client_form.html', {'form': form})

@login_required
def client_update(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientForm(instance=client)
    return render(request, 'backend/client_form.html', {'form': form, 'client': client})

@login_required
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.delete()
        return redirect('client_list')
    return render(request, 'backend/client_confirm_delete.html', {'client': client})

@login_required
def invoice_list(request):
    invoices = Invoice.objects.all()

    if request.method == 'POST' and 'export' in request.POST:
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define the column headers
        column_headers = ['Contractor Name', 'Invoice Number', 'Invoice Date', 'Internal Invoice Number', 'Internal Invoice Date']

        # Write the column headers to the worksheet
        for i, header in enumerate(column_headers, start=1):
            ws[f'A{i}'] = header

        # Write the data to the worksheet
        for i, invoice in enumerate(invoices, start=2):
            ws[f'B{i}'] = invoice.invoice_number
            ws[f'C{i}'] = invoice.invoice_date
            ws[f'D{i}'] = invoice.internal_invoice_number
            ws[f'E{i}'] = invoice.internal_invoice_date

        # Create an HTTP response with the workbook as the content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=invoices.xlsx'
        wb.save(response)
        return response

    return render(request, 'backend/invoice_list.html', {'invoices': invoices})

@login_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('invoice_list')
    else:
        form = InvoiceForm()
    return render(request, 'backend/invoice_form.html', {'form': form})

@login_required
def invoice_update(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('invoice_list')
    else:
        form = InvoiceForm(instance=invoice)
    return render(request, 'backend/invoice_form.html', {'form': form, 'invoice': invoice})

@login_required
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        invoice.delete()
        return redirect('invoice_list')
    return render(request, 'backend/invoice_confirm_delete.html', {'invoice': invoice})

@login_required
@login_required
def expense_list(request):
    expenses = Expense.objects.all()

    if request.method == 'POST' and 'export' in request.POST:
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define the column headers
        column_headers = ['Contractor Name', 'Expense', 'Profit / Loss']

        # Write the column headers to the worksheet
        for i, header in enumerate(column_headers, start=1):
            ws[f'A{i}'] = header

        # Write the data to the worksheet
        for i, expense in enumerate(expenses, start=2):
            ws[f'A{i}'] = expense.contract.practice_name.practice_name
            ws[f'B{i}'] = expense.expenses
            ws[f'C{i}'] = expense.profit_loss

        # Create an HTTP response with the workbook as the content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
        wb.save(response)
        return response

    return render(request, 'backend/expense_list.html', {'expenses': expenses})

@login_required
def expense_create(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'backend/expense_form.html', {'form': form})

@login_required
def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect('expense_list')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'backend/expense_form.html', {'form': form, 'expense': expense})

@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.delete()
        return redirect('expense_list')
    return render(request, 'backend/expense_confirm_delete.html', {'expense': expense})

@login_required
def party_list(request):
    parties = Party.objects.all()

    if request.method == 'POST' and 'export' in request.POST:
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define the column headers
        column_headers = ['Party-1 Name', 'Party-1 Amount', 'Party-2 Name', 'Party-2 Amount']

        # Write the column headers to the worksheet
        for i, header in enumerate(column_headers, start=1):
            ws[f'A{i}'] = header

        # Write the data to the worksheet
        for i, party in enumerate(parties, start=2):
            ws[f'A{i}'] = party.party_1
            ws[f'B{i}'] = party.party_1_amount
            ws[f'C{i}'] = party.party_2
            ws[f'D{i}'] = party.party_2_amount

        # Create an HTTP response with the workbook as the content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=parties.xlsx'
        wb.save(response)
        return response

    return render(request, 'backend/party_list.html', {'parties': parties})
@login_required
def party_create(request):
    if request.method == 'POST':
        form = PartyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('party_list')
    else:
        form = PartyForm()
    return render(request, 'backend/party_form.html', {'form': form})

@login_required
def party_update(request, pk):
    party = get_object_or_404(Party, pk=pk)
    if request.method == 'POST':
        form = PartyForm(request.POST, instance=party)
        if form.is_valid():
            form.save()
            return redirect('party_list')
    else:
        form = PartyForm(instance=party)
    return render(request, 'backend/party_form.html', {'form': form, 'party': party})

@login_required
def party_delete(request, pk):
    party = get_object_or_404(Party, pk=pk)
    if request.method == 'POST':
        party.delete()
        return redirect('party_list')
    return render(request, 'backend/party_confirm_delete.html', {'party': party})

@login_required
def pending_invoices(request):
    pending_invoices = Contract.objects.filter(invoice_date__isnull=True)
    context = {'pending_invoices': pending_invoices}
    return render(request, 'backend/pending_invoices.html', context)

@login_required
def generate_reports(request):
    return render(request, 'backend/generate_reports.html')
